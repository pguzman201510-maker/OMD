
import pandas as pd
import os
from datetime import datetime, date
import io
import openpyxl

class DataManager:
    def __init__(self, uvr_file, inflation_file, consolidado_file):
        self.uvr_file = uvr_file
        self.inflation_file = inflation_file
        self.consolidado_file = consolidado_file

    def get_uvr(self, query_date):
        """Returns UVR value for a specific date."""
        try:
            if isinstance(self.uvr_file, str):
                df = pd.read_excel(self.uvr_file)
            else:
                self.uvr_file.seek(0)
                df = pd.read_excel(self.uvr_file)

            df.iloc[:, 0] = pd.to_datetime(df.iloc[:, 0], errors='coerce')

            row = df[df.iloc[:, 0] == pd.to_datetime(query_date)]
            if not row.empty:
                return float(row.iloc[0, 1])
            return 100.0 # Default fallback
        except Exception as e:
            print(f"Error reading UVR: {e}")
            return 100.0

    def get_inflation(self, year):
        """Returns annual inflation for the given year."""
        try:
            if isinstance(self.inflation_file, str):
                df = pd.read_excel(self.inflation_file)
            else:
                self.inflation_file.seek(0)
                df = pd.read_excel(self.inflation_file)

            return float(df.iloc[-1, 2]) / 100.0
        except:
            return 0.03 # 3% Default

    def save_results(self, details_df, summary_stats):
        """
        Updates the Excel file with the Stacked Matrix format.
        Target Sheets: 'Tesorería' or 'Mercado' based on operation type.
        Four tables to update per sheet.
        """
        op_type = summary_stats.get('op_type', 'Mercado') # Tesorería / Mercado
        target_sheet = op_type if op_type in ["Tesorería", "Mercado"] else "Mercado"

        # Load workbook
        if isinstance(self.consolidado_file, str):
            wb = openpyxl.load_workbook(self.consolidado_file)
        else:
            self.consolidado_file.seek(0)
            wb = openpyxl.load_workbook(self.consolidado_file)

        # Ensure sheet exists
        if target_sheet not in wb.sheetnames:
            wb.create_sheet(target_sheet)

        ws = wb[target_sheet]

        # If sheet is empty, create from scratch
        if ws.max_row <= 1:
            self._init_sheet_structure(ws)

        settlement_date = summary_stats.get('fecha_liq')
        date_col_header = settlement_date.strftime('%Y-%m-%d')

        # Split details_df into Rec/Ent
        df_rec = details_df[details_df["Tipo"] == "Recogido"]
        df_ent = details_df[details_df["Tipo"] == "Entregado"]

        # Iterative Update: Find Start -> Update -> Return New End (Shift) -> Repeat
        # Order is known:
        # 1. Recibidos Nominal
        # 2. Recibidos Costo
        # 3. Entregados Nominal
        # 4. Entregados Costo

        current_search_start = 0

        # 1. Recibidos Nominal
        start_row = self._find_table_header(ws, "TITULOS RECIBIDOS (NOMINAL)", current_search_start)
        if start_row is not None:
            _, rows_added = self._update_table_block(ws, start_row, df_rec, "Nominal Orig", date_col_header)
            current_search_start = start_row + rows_added + 1

        # 2. Recibidos Costo
        start_row = self._find_table_header(ws, "VALOR COSTO COP (RECIBIDOS)", current_search_start)
        if start_row is not None:
            _, rows_added = self._update_table_block(ws, start_row, df_rec, "Valor Costo", date_col_header)
            current_search_start = start_row + rows_added + 1

        # 3. Entregados Nominal
        start_row = self._find_table_header(ws, "TITULOS ENTREGADOS (NOMINAL)", current_search_start)
        if start_row is not None:
            _, rows_added = self._update_table_block(ws, start_row, df_ent, "Nominal Orig", date_col_header)
            current_search_start = start_row + rows_added + 1

        # 4. Entregados Costo
        start_row = self._find_table_header(ws, "VALOR COSTO COP (ENTREGADOS)", current_search_start)
        if start_row is not None:
            _, rows_added = self._update_table_block(ws, start_row, df_ent, "Valor Costo", date_col_header)

        # Save
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _init_sheet_structure(self, ws):
        """Initializes an empty sheet with the 4 table headers."""
        ws.append(["TITULOS RECIBIDOS (NOMINAL)"])
        ws.append(["ISIN", "VENCIMIENTO", "MONEDA", "CUPON"])
        ws.append([]) # Gap

        ws.append(["VALOR COSTO COP (RECIBIDOS)"])
        ws.append(["ISIN", "VENCIMIENTO", "MONEDA", "CUPON"])
        ws.append([])

        ws.append(["TITULOS ENTREGADOS (NOMINAL)"])
        ws.append(["ISIN", "VENCIMIENTO", "MONEDA", "CUPON"])
        ws.append([])

        ws.append(["VALOR COSTO COP (ENTREGADOS)"])
        ws.append(["ISIN", "VENCIMIENTO", "MONEDA", "CUPON"])
        ws.append([])

    def _find_table_header(self, ws, keyword, start_offset=0):
        """Finds row index of a cell containing keyword, starting search at start_offset."""
        for i, row in enumerate(ws.iter_rows(min_row=start_offset+1, values_only=True)):
            # iter_rows 1-based, min_row argument. 'i' is 0-based relative to min_row.
            # Real row index = start_offset + i
            if row[0] and isinstance(row[0], str) and keyword in row[0].upper():
                return start_offset + i
        return None

    def _update_table_block(self, ws, start_row_idx, new_data_df, value_col, date_header):
        """
        Updates a specific table block in the worksheet.
        start_row_idx: Index of the Title Row (0-based)
        Returns: (new_end_row_idx, rows_added_count)
        """
        header_row_idx = start_row_idx + 1 # 0-based

        # Helper to read headers from a specific row index (0-based)
        def get_headers(r_idx):
            return [c.value for c in list(ws.rows)[r_idx] if c.value is not None]

        # Since we might insert columns, reading 'headers' once is fine, as col insertion doesn't shift rows.
        # But we must be careful with 'ws.rows' generator if we modify structure.
        # Safest to access by coordinate.

        # Read Headers
        headers = []
        c = 1
        while True:
            val = ws.cell(row=header_row_idx+1, column=c).value
            if val is None and c > 4: # Stop if empty after fixed cols
                break
            if val is not None:
                headers.append(val)
            c += 1

        # Check if Date Header exists
        if date_header not in headers:
            new_col_idx = len(headers) + 1
            ws.cell(row=header_row_idx+1, column=new_col_idx, value=date_header)
            headers.append(date_header)

        date_col_idx = headers.index(date_header) # 0-based in headers list

        # Map ISINs
        isin_map = {}
        rows_added = 0

        # Scan for existing data rows until next empty row or Header
        current_row = header_row_idx + 1 # 0-based

        while True:
            # Safe check bound
            if current_row >= ws.max_row:
                break

            cell_val = ws.cell(row=current_row+1, column=1).value
            if cell_val is None:
                break # Gap detected

            s_val = str(cell_val).strip()
            if "TITULOS" in s_val or "VALOR COSTO" in s_val:
                break # Next table hit

            isin_map[s_val] = current_row
            current_row += 1

        # current_row is now the index where we would insert (the gap)
        insert_pos = current_row + 1 # 1-based for openpyxl

        for _, row in new_data_df.iterrows():
            isin = str(row.get("ISIN", "")).strip()
            if not isin: continue

            if isin in isin_map:
                target_row = isin_map[isin] + 1 # 1-based
            else:
                # Insert
                ws.insert_rows(insert_pos)
                target_row = insert_pos

                ws.cell(row=target_row, column=1, value=isin)
                ws.cell(row=target_row, column=2, value=str(row.get("Vencimiento", "")))
                ws.cell(row=target_row, column=3, value=row.get("Denom (COP/UVR)", "COP"))
                ws.cell(row=target_row, column=4, value=row.get("Cupón %", 0.0))

                isin_map[isin] = insert_pos - 1
                insert_pos += 1
                rows_added += 1

            val = row.get(value_col, 0.0)
            ws.cell(row=target_row, column=date_col_idx + 1, value=val)

        return (insert_pos, rows_added)
